#!/usr/bin/env python3
"""
BSA Opportunity Scout — main.py
Full pipeline: discover sources → scrape → parse PDFs → assess fit → build Excel → email
Run: python scripts/main.py
"""

import functools
import io
import json
import logging
import os
import re
import smtplib
import time
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from urllib.parse import urljoin

import anthropic
import openpyxl
import pdfplumber
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ── Config ───────────────────────────────────────────────────────────────────
COMPANY_PROFILE = Path(__file__).parent.parent / "knowledgebase" / "company_profile.txt"
OUTPUT_DIR      = Path(__file__).parent.parent / "output"
RECIPIENT       = "alex@bsa.ai"
SENDER_EMAIL    = os.getenv("SENDER_EMAIL")
SENDER_PASS     = os.getenv("SENDER_APP_PASSWORD")
MODEL           = "claude-sonnet-4-6"

_api_key = os.getenv("ANTHROPIC_API_KEY")
if not _api_key:
    log.warning("ANTHROPIC_API_KEY not set — Claude features will fail")
CLAUDE_CLIENT = anthropic.Anthropic(
    api_key=_api_key or "missing",
    timeout=60.0,        # seconds per request
    max_retries=2,       # Anthropic SDK-level retries before raising
)

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; BSA-Scout/1.0; +https://bsa.ai)"}
TIMEOUT = 30

KEYWORDS = [
    "research", "data", "analytics", "finance", "credit", "digital", "facilitation",
    "consulting", "advisory", "technical assistance", "fintech", "financial inclusion",
    "capacity building", "monitoring", "evaluation", "ict", "information systems",
    "software", "database", "survey", "assessment", "study", "feasibility",
    "artificial intelligence", "machine learning", "blockchain", "risk management",
]

# JS-heavy domains — use Playwright
_JS_DOMAINS = ["worldbank", "europa.eu", "ungm", "afdb.org", "undp", "unicef", "ausschreibungen"]

SITES: list[dict] = [
    {"name": "UNGM",         "url": "https://www.ungm.org/Public/Notice",                                                                                    "country_filter": "Tanzania", "use_playwright": True},
    {"name": "World Bank",   "url": "https://projects.worldbank.org/en/projects-operations/opportunities?srce=both&project_ctry_name_exact=Tanzania",         "country_filter": None,       "use_playwright": True},
    {"name": "EU Tenders",   "url": "https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/calls-for-tenders?isExactMatch=true&order=DESC&pageNumber=1&pageSize=50&sortBy=startDate", "country_filter": "Tanzania", "use_playwright": True},
    {"name": "US Embassy TZ","url": "https://tz.usembassy.gov/contract-opportunities/",                                                                       "country_filter": None,       "use_playwright": False},
    {"name": "AfDB",         "url": "https://www.afdb.org/en/projects-and-operations/procurement",                                                            "country_filter": "Tanzania", "use_playwright": True},
    {"name": "GIZ Tanzania", "url": "https://www.giz.de/en/regions/africa/tanzania/tenders",                                                                  "country_filter": None,       "use_playwright": False},
    {"name": "DG Market",    "url": "https://www.dgmarket.com/tenders/list.do?sub=services-in-Tanzania-2&locationISO=tz",                                     "country_filter": None,       "use_playwright": False},
    {"name": "FSDT",         "url": "https://www.fsdt.or.tz/work-with-us/",                                                                                   "country_filter": None,       "use_playwright": False},
    {"name": "CRDB Bank",    "url": "https://crdbbank.co.tz/en/about-us/tender",                                                                              "country_filter": None,       "use_playwright": False},
    {"name": "NBC Bank",     "url": "https://www.nbc.co.tz/en/procurement/",                                                                                  "country_filter": None,       "use_playwright": False},
    {"name": "NMB Bank",     "url": "https://www.nmbbank.co.tz/tenders",                                                                                      "country_filter": None,       "use_playwright": False},
    {"name": "GIZ Ausschr.", "url": "https://ausschreibungen.giz.de/Satellite/company/welcome.do",                                                            "country_filter": "Tanzania", "use_playwright": True},
    {"name": "India HC TZ",  "url": "https://hcindiatz.gov.in/tenders-tanzania.php",                                                                         "country_filter": None,       "use_playwright": False},
]

# ── Retry decorator ───────────────────────────────────────────────────────────
def retry(max_retries: int = 3, backoff: float = 2.0, exceptions: tuple = (Exception,)):
    """Retry a function with exponential backoff on specified exceptions."""
    def decorator(fn):
        @functools.wraps(fn)
        def wrapper(*args, **kwargs):
            last_exc: Exception | None = None
            for attempt in range(max_retries):
                try:
                    return fn(*args, **kwargs)
                except exceptions as exc:
                    last_exc = exc
                    if attempt < max_retries - 1:
                        wait = backoff ** attempt
                        log.warning(
                            f"{fn.__name__} attempt {attempt + 1}/{max_retries} failed: {exc!r}. "
                            f"Retrying in {wait:.1f}s..."
                        )
                        time.sleep(wait)
            raise last_exc  # type: ignore[misc]
        return wrapper
    return decorator


# ── Helpers ───────────────────────────────────────────────────────────────────
def load_profile() -> str:
    if not COMPANY_PROFILE.exists():
        raise FileNotFoundError(
            f"company_profile.txt not found at {COMPANY_PROFILE.resolve()}. "
            "Ensure knowledgebase/company_profile.txt exists."
        )
    text = COMPANY_PROFILE.read_text(encoding="utf-8").strip()
    if not text:
        raise ValueError("company_profile.txt is empty.")
    return text


def is_relevant(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in KEYWORDS)


@retry(max_retries=3, backoff=2.0, exceptions=(requests.RequestException, OSError))
def _safe_get_retried(url: str) -> requests.Response | None:
    """
    Inner HTTP GET — retried up to 3× on transient failures.
    Returns None for permanent 4xx errors (no retry). Raises on 5xx / timeouts
    so the retry decorator can catch and retry them.
    """
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
        return r
    except requests.HTTPError as e:
        status = getattr(getattr(e, "response", None), "status_code", None)
        if status in (403, 404, 410):
            log.warning(f"Permanent HTTP {status} for {url} — skipping (no retry)")
            return None  # return None → retry decorator sees success, stops retrying
        raise  # 5xx → retry decorator catches and retries


def safe_get(url: str) -> requests.Response | None:
    """
    HTTP GET with automatic retry. Always returns None on failure — never raises.
    Callers can treat None as "this URL is unavailable".
    """
    try:
        return _safe_get_retried(url)
    except Exception as e:
        log.warning(f"GET failed for {url} after retries: {e!r}")
        return None


def find_pdf_links(soup: BeautifulSoup, base_url: str) -> list[str]:
    pdfs = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.lower().endswith(".pdf"):
            continue
        # Use urljoin to properly handle absolute, root-relative, and path-relative URLs
        full = urljoin(base_url.rstrip("/") + "/", href)
        pdfs.append(full)
    return pdfs[:3]  # cap at 3 PDFs per page


def parse_pdf(url: str) -> str:
    """Download and extract text from a PDF. Returns empty string on any failure."""
    r = safe_get(url)
    if not r:
        return ""
    try:
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        return "\n".join(pages_text)[:6000]
    except Exception as e:
        log.warning(f"PDF parse failed for {url}: {e}")
        return ""


# ── Scraper ───────────────────────────────────────────────────────────────────
def _passes_country_filter(site: dict, anchor, href: str) -> bool:
    cf = site.get("country_filter")
    if not cf:
        return True
    parent_text = anchor.parent.get_text(" ", strip=True).lower() if anchor.parent else ""
    return cf.lower() in parent_text or cf.lower() in href.lower()


def _fetch_detail(full_url: str) -> tuple[str, str]:
    """Return (detail_text, pdf_text) for a detail/opportunity page."""
    detail_r = safe_get(full_url)
    if not detail_r:
        return "", ""
    detail_soup = BeautifulSoup(detail_r.text, "html.parser")
    detail_text = detail_soup.get_text(" ", strip=True)[:4000]
    pdf_text    = "".join(parse_pdf(u) for u in find_pdf_links(detail_soup, full_url))
    return detail_text, pdf_text


def _get_page_html(site: dict) -> str:
    """
    Fetch page HTML. Uses Playwright for JS-heavy sites (use_playwright=True),
    falling back to requests if Playwright is unavailable or fails.
    """
    if PLAYWRIGHT_AVAILABLE and site.get("use_playwright"):
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page(extra_http_headers=HEADERS)
                page.goto(site["url"], wait_until="networkidle", timeout=30_000)
                html = page.content()
                browser.close()
                return html
        except Exception as e:
            log.warning(
                f"Playwright failed for {site['name']}: {e}. "
                "Falling back to requests (may miss JS-rendered content)."
            )
    r = safe_get(site["url"])
    return r.text if r else ""


def scrape_site(site: dict) -> list[dict]:
    """Scrape a single portal, returning relevant opportunity dicts."""
    log.info(f"Scraping {site['name']} ...")
    html = _get_page_html(site)
    if not html:
        log.warning(f"No HTML retrieved from {site['name']} — skipping")
        return []

    soup  = BeautifulSoup(html, "html.parser")
    found: list[dict] = []
    seen:  set[str]   = set()

    for anchor in soup.find_all("a", href=True):
        title = anchor.get_text(strip=True)
        href  = anchor["href"]

        if not title or len(title) < 10 or href in seen:
            continue
        if not is_relevant(title):
            continue
        if not _passes_country_filter(site, anchor, href):
            continue

        full_url = href if href.startswith("http") else site["url"].rstrip("/") + "/" + href.lstrip("/")
        seen.add(href)

        detail_text, pdf_text = _fetch_detail(full_url)
        if not detail_text:
            detail_text = title

        raw_text = (detail_text + "\n\nPDF CONTENT:\n" + pdf_text) if pdf_text else detail_text
        found.append({
            "source":     site["name"],
            "source_url": full_url,
            "title":      title,
            "raw_text":   raw_text,
        })
        time.sleep(0.8)

    log.info(f"  → {len(found)} relevant items from {site['name']}")
    return found


# ── Claude source discovery ───────────────────────────────────────────────────
DISCOVER_PROMPT = """\
You are helping a Tanzanian data analytics and fintech consulting firm find procurement portals.

Company profile summary:
{profile}

Search the web for procurement portals, tender pages, and consulting opportunity listings
specifically covering Tanzania in these service areas:
- Data analytics, research, AI/machine learning
- Finance, fintech, credit scoring
- Capacity building, facilitation, training
- Monitoring & evaluation, digital transformation
- ICT and information systems

Find direct URLs to the tenders/opportunities listing pages for:
1. UN agency Tanzania procurement portals (UNDP, UNICEF, WHO, FAO, ILO, UN Women, UNOPS)
2. Tanzania government portals (PPRA, Ministry of Finance, TANESCO, TRA, NBS)
3. Bilateral donors (KfW, SIDA, FCDO, USAID, GIZ sub-portals, Danida, JICA)
4. East Africa development banks and foundations (Aga Khan Foundation, MasterCard Foundation)
5. Private sector banks and companies in Tanzania posting tenders

Return ONLY a valid JSON array — no explanation, no markdown:
[{{"name": "Portal Name", "url": "https://direct-url-to-tenders", "country_filter": null}}]
"""


def discover_sources(profile: str, existing_urls: set[str]) -> list[dict]:
    """
    Use Claude with web search to find new Tanzania consulting opportunity portals.
    Returns a list of new site dicts (deduped against existing_urls).
    Gracefully returns [] on any failure.
    """
    log.info("Asking Claude to discover new opportunity sources via web search...")
    try:
        resp = CLAUDE_CLIENT.messages.create(
            model=MODEL,
            max_tokens=2000,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 8}],
            messages=[{
                "role": "user",
                "content": DISCOVER_PROMPT.format(profile=profile[:800]),
            }],
        )

        # Extract the last text block (Claude's final synthesis after web searches)
        text = ""
        for block in reversed(resp.content):
            if hasattr(block, "text") and block.text.strip():
                text = block.text.strip()
                break

        if not text:
            log.warning("discover_sources: Claude returned no text content")
            return []

        # Strip markdown fences, then find JSON array
        text = re.sub(r"(^```json\s*)|(\s*```$)", "", text, flags=re.MULTILINE).strip()
        match = re.search(r"\[[^\]]*(?:\[[^\]]*\][^\]]*)*\]", text, re.DOTALL)
        if not match:
            log.warning("discover_sources: No JSON array found in Claude response")
            return []

        raw_sites: list = json.loads(match.group())

        new_sites = []
        for s in raw_sites:
            if not isinstance(s, dict) or not s.get("url"):
                continue
            if s["url"] in existing_urls:
                continue
            s.setdefault("country_filter", None)
            s.setdefault("use_playwright", any(d in s["url"] for d in _JS_DOMAINS))
            new_sites.append(s)

        log.info(f"Claude discovered {len(new_sites)} new opportunity sources")
        return new_sites

    except Exception as e:
        log.warning(f"discover_sources failed ({e!r}) — continuing with known sites only")
        return []


# ── Assessor ─────────────────────────────────────────────────────────────────
ASSESS_PROMPT = """\
You are a business development analyst for a Tanzanian consulting firm.
Assess whether the following opportunity is a good fit for the company.

COMPANY PROFILE:
{profile}

OPPORTUNITY DETAILS (title, webpage text, any PDF content):
{opp_text}

SOURCE WEBSITE: {source}

Return ONLY a valid JSON object with these exact keys:
{{
  "title": "short clean title",
  "organization": "issuing org name",
  "type": "RFP | EOI | Call for Tenders | Grant | Expression of Interest | Other",
  "sectors": ["sector1","sector2"],
  "deadline": "DD Month YYYY or Not specified",
  "budget": "USD X,XXX,XXX or Not specified",
  "budget_type": "fixed | to_be_proposed | not_specified",
  "eligibility": ["criterion 1","criterion 2","criterion 3"],
  "fit_assessment": "fit | nearly_fit | far_fetched",
  "fit_reasoning": "2-sentence explanation of why",
  "contact_email": "email@domain.com or Not found",
  "application_link": "https://... or Not found",
  "background_context": "2-3 sentences describing what this call is about and who issued it"
}}
No markdown, no explanation, only JSON.
"""

_REQUIRED_KEYS = {
    "title", "organization", "type", "sectors", "deadline", "budget",
    "budget_type", "eligibility", "fit_assessment", "fit_reasoning",
    "contact_email", "application_link", "background_context",
}
_VALID_FIT = {"fit", "nearly_fit", "far_fetched"}


def _validate_assessment(data: dict) -> bool:
    """Return True if the Claude JSON response has all required keys and valid values."""
    missing = _REQUIRED_KEYS - set(data.keys())
    if missing:
        log.warning(f"Assessment missing keys: {missing}")
        return False
    if data.get("fit_assessment") not in _VALID_FIT:
        log.warning(f"Invalid fit_assessment value: '{data.get('fit_assessment')}'")
        return False
    if not isinstance(data.get("sectors"), list):
        log.warning("Assessment 'sectors' is not a list — coercing")
        data["sectors"] = [str(data.get("sectors", ""))]
    if not isinstance(data.get("eligibility"), list):
        log.warning("Assessment 'eligibility' is not a list — coercing")
        data["eligibility"] = [str(data.get("eligibility", ""))]
    return True


def assess(opp: dict, profile: str) -> dict | None:
    """Ask Claude to assess fit for one opportunity. Returns None on failure."""
    prompt = ASSESS_PROMPT.format(
        profile=profile[:2000],
        opp_text=opp["raw_text"][:4000],
        source=opp["source"],
    )
    try:
        resp = CLAUDE_CLIENT.messages.create(
            model=MODEL,
            max_tokens=900,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        text = re.sub(r"(^```json\s*)|(\s*```$)", "", text, flags=re.MULTILINE)
        data = json.loads(text)

        if not _validate_assessment(data):
            log.warning(f"Skipping malformed assessment for: {opp['title'][:60]}")
            return None

        data["source_url"] = opp["source_url"]
        data["source"]     = opp["source"]
        return data

    except json.JSONDecodeError as e:
        log.warning(f"Claude returned invalid JSON for '{opp['title'][:60]}': {e}")
        return None
    except anthropic.APIStatusError as e:
        log.error(f"Claude API error (status {e.status_code}) for '{opp['title'][:60]}': {e.message}")
        return None
    except anthropic.APIConnectionError as e:
        log.error(f"Claude API connection error for '{opp['title'][:60]}': {e}")
        return None
    except Exception as e:
        log.warning(f"Assessment failed for '{opp['title'][:60]}': {e!r}")
        return None


# ── Excel Builder ─────────────────────────────────────────────────────────────
COLS = [
    ("Title",            38), ("Organization",   22), ("Type",             14),
    ("Sectors",          22), ("Deadline",        16), ("Budget",           18),
    ("Budget Type",      14), ("Fit Assessment",  14), ("Fit Reasoning",    38),
    ("Eligibility",      40), ("Contact Email",   26), ("Application Link", 30),
    ("Background",       48), ("Source",          14), ("Source URL",       30),
    ("Date Found",       14),
]

FIT_COLORS = {"fit": "C6EFCE", "nearly_fit": "FFEB9C", "far_fetched": "FFC7CE"}
HEADER_BG  = "1F4E79"


def build_excel(opps: list[dict], path: Path) -> None:
    """Build a colour-coded Excel workbook with an Opportunities sheet and Summary sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opportunities"
    ws.freeze_panes = "A2"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (hdr, w) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.fill      = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        cell.font      = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28

    for r, o in enumerate(opps, 2):
        fit   = o.get("fit_assessment", "").lower()
        color = FIT_COLORS.get(fit, "FFFFFF")
        fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")

        vals = [
            o.get("title", ""),
            o.get("organization", ""),
            o.get("type", ""),
            ", ".join(o.get("sectors", [])),
            o.get("deadline", ""),
            o.get("budget", ""),
            o.get("budget_type", "").replace("_", " ").title(),
            o.get("fit_assessment", "").replace("_", " ").upper(),
            o.get("fit_reasoning", ""),
            "\n".join(f"• {e}" for e in o.get("eligibility", [])),
            o.get("contact_email", ""),
            o.get("application_link", ""),
            o.get("background_context", ""),
            o.get("source", ""),
            o.get("source_url", ""),
            datetime.now().strftime("%Y-%m-%d"),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2    = wb.create_sheet("Summary")
    fit_c  = sum(1 for o in opps if o.get("fit_assessment") == "fit")
    near_c = sum(1 for o in opps if o.get("fit_assessment") == "nearly_fit")
    far_c  = sum(1 for o in opps if o.get("fit_assessment") == "far_fetched")

    ws2.append(["BSA Weekly Opportunity Report", datetime.now().strftime("%d %B %Y")])
    ws2.append([])
    ws2.append(["Total Opportunities", len(opps)])
    ws2.append(["Strong Fit",          fit_c])
    ws2.append(["Nearly Fit",          near_c])
    ws2.append(["Far Fetched",         far_c])
    ws2.append([])
    ws2.append(["TOP STRONG FIT OPPORTUNITIES"])
    for o in [x for x in opps if x.get("fit_assessment") == "fit"][:8]:
        ws2.append([o.get("title", ""), o.get("organization", ""), o.get("deadline", ""), o.get("budget", "")])

    wb.save(path)
    log.info(f"Excel saved → {path}")


# ── Emailer ───────────────────────────────────────────────────────────────────
def _build_email_body(opps: list[dict]) -> tuple[str, int, int, int]:
    """Build the plain-text email body. Returns (body, fit_count, near_count, far_count)."""
    fit_c  = sum(1 for o in opps if o.get("fit_assessment") == "fit")
    near_c = sum(1 for o in opps if o.get("fit_assessment") == "nearly_fit")
    far_c  = sum(1 for o in opps if o.get("fit_assessment") == "far_fetched")

    top_lines = "\n".join(
        f"  • {o.get('title','N/A')} | {o.get('organization','N/A')} "
        f"| Deadline: {o.get('deadline','N/A')} | Budget: {o.get('budget','N/A')}"
        for o in [x for x in opps if x.get("fit_assessment") == "fit"][:5]
    )

    body = (
        f"Hi Alex,\n\n"
        f"Your weekly Tanzania consulting opportunity scan is complete.\n\n"
        f"{'━' * 38}\n"
        f" WEEKLY SUMMARY — {datetime.now().strftime('%d %B %Y')}\n"
        f"{'━' * 38}\n"
        f" Total scanned : {len(opps)}\n"
        f" Strong Fit    : {fit_c}\n"
        f" Nearly Fit    : {near_c}\n"
        f" Far Fetched   : {far_c}\n\n"
        f"TOP STRONG FIT OPPORTUNITIES:\n"
        f"{top_lines if top_lines else '  None found this week.'}\n\n"
        f"Full details in the attached Excel report (colour-coded by fit).\n\n"
        f"Best regards,\nBSA Opportunity Scout\n"
    )
    return body, fit_c, near_c, far_c


def _save_email_fallback(body: str, excel_path: Path) -> None:
    """If email fails, save the body text next to the Excel file for manual review."""
    fallback = excel_path.with_suffix(".email_body.txt")
    fallback.write_text(body, encoding="utf-8")
    log.info(f"Email body saved to fallback file → {fallback}")


def send_email(excel_path: Path, opps: list[dict]) -> bool:
    """
    Send the weekly report email with Excel attachment.
    Returns True on success, False on failure (failure is logged, not raised).
    """
    if not SENDER_EMAIL or not SENDER_PASS:
        log.warning(
            "Email credentials not configured (SENDER_EMAIL / SENDER_APP_PASSWORD). "
            f"Skipping email — Excel saved at {excel_path}"
        )
        return False

    body, fit_c, _, _ = _build_email_body(opps)

    msg           = MIMEMultipart()
    msg["From"]   = SENDER_EMAIL
    msg["To"]     = RECIPIENT
    # Keep subject ASCII to avoid RFC2047 encoding in raw message headers (tests parse raw lines).
    msg["Subject"] = (
        f"BSA Weekly Opportunities - {datetime.now().strftime('%d %b %Y')} "
        f"({fit_c} Strong Fit)"
    )
    msg.attach(MIMEText(body, "plain"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{excel_path.name}"')
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(SENDER_EMAIL, SENDER_PASS)
            s.sendmail(SENDER_EMAIL, RECIPIENT, msg.as_string())
        log.info(f"Email sent to {RECIPIENT}")
        return True

    except smtplib.SMTPAuthenticationError:
        log.error(
            "SMTP authentication failed. Check SENDER_EMAIL and SENDER_APP_PASSWORD. "
            "For Gmail/Google Workspace, use an App Password (not your account password). "
            "Enable 2FA first, then generate at myaccount.google.com/apppasswords"
        )
        _save_email_fallback(body, excel_path)
        return False

    except smtplib.SMTPRecipientsRefused as e:
        log.error(f"Recipient refused by SMTP server: {e.recipients}")
        _save_email_fallback(body, excel_path)
        return False

    except smtplib.SMTPException as e:
        log.error(f"SMTP error while sending email: {e}")
        _save_email_fallback(body, excel_path)
        return False

    except OSError as e:
        log.error(f"Network error connecting to SMTP server: {e}")
        _save_email_fallback(body, excel_path)
        return False


# ── Main pipeline ─────────────────────────────────────────────────────────────
def main() -> int:
    """
    Full pipeline. Returns exit code: 0 = success, 1 = no opportunities found,
    2 = profile load failure.
    """
    log.info("=== BSA Opportunity Scout starting ===")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load company profile
    try:
        profile = load_profile()
        log.info("Company profile loaded.")
    except (FileNotFoundError, ValueError) as e:
        log.error(f"Cannot load company profile: {e}")
        return 2

    # Step 1 — Discover new sources with Claude web search
    existing_urls = {s["url"] for s in SITES}
    new_sites     = discover_sources(profile, existing_urls)
    all_sites     = SITES + new_sites
    if new_sites:
        log.info(f"Added {len(new_sites)} Claude-discovered sources: {[s['name'] for s in new_sites]}")

    # Step 2 — Scrape all sites
    raw: list[dict] = []
    for site in all_sites:
        try:
            raw.extend(scrape_site(site))
        except Exception as e:
            log.error(f"Site '{site['name']}' failed unexpectedly: {e!r} — continuing")

    log.info(f"Total raw opportunities found: {len(raw)}")
    if not raw:
        log.warning("No opportunities scraped. Check site connectivity and keyword list.")

    # Step 3 — Deduplicate by normalised title
    seen_titles: set[str] = set()
    deduped: list[dict]   = []
    for o in raw:
        key = re.sub(r"\W+", " ", o["title"].lower()).strip()[:60]
        if key not in seen_titles:
            seen_titles.add(key)
            deduped.append(o)
    log.info(f"After deduplication: {len(deduped)} opportunities")

    # Step 4 — Assess each with Claude
    assessed: list[dict] = []
    for i, opp in enumerate(deduped, 1):
        log.info(f"Assessing {i}/{len(deduped)}: {opp['title'][:60]}")
        result = assess(opp, profile)
        if result:
            assessed.append(result)
        time.sleep(1.2)  # respect Claude rate limits

    if not assessed:
        log.warning("No opportunities successfully assessed.")
        return 1

    order = {"fit": 0, "nearly_fit": 1, "far_fetched": 2}
    assessed.sort(key=lambda x: order.get(x.get("fit_assessment", ""), 3))

    # Step 5 — Build Excel
    date_str   = datetime.now().strftime("%Y%m%d")
    excel_path = OUTPUT_DIR / f"BSA_Opportunities_{date_str}.xlsx"
    build_excel(assessed, excel_path)

    # Step 6 — Send email
    send_email(excel_path, assessed)

    fit_c = sum(1 for o in assessed if o.get("fit_assessment") == "fit")
    log.info(
        f"=== Done === | {len(assessed)} total | {fit_c} strong fit | "
        f"Excel: {excel_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
