"""
Tests for build_excel(): correct sheets, column headers, row colours, summary counts.
"""
import openpyxl
import pytest

import main as m


def _make_opp(fit: str, title: str = "Test Opportunity") -> dict:
    return {
        "title":            title,
        "organization":     "Test Org",
        "type":             "RFP",
        "sectors":          ["Finance", "Data"],
        "deadline":         "30 June 2026",
        "budget":           "USD 50,000",
        "budget_type":      "fixed",
        "fit_assessment":   fit,
        "fit_reasoning":    "Good match.",
        "eligibility":      ["Registered firm", "5+ years"],
        "contact_email":    "test@org.com",
        "application_link": "https://org.com/apply",
        "background_context": "Context here.",
        "source":           "Test",
        "source_url":       "https://test.com",
    }


class TestBuildExcel:
    def test_creates_file(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([_make_opp("fit")], path)
        assert path.exists()

    def test_has_two_sheets(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([_make_opp("fit")], path)
        wb = openpyxl.load_workbook(path)
        assert set(wb.sheetnames) == {"Opportunities", "Summary"}

    def test_header_row_has_all_columns(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([], path)
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Opportunities"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(m.COLS) + 1)]
        expected = [col[0] for col in m.COLS]
        assert headers == expected

    def test_fit_row_is_green(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([_make_opp("fit")], path)
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Opportunities"]
        cell = ws.cell(row=2, column=1)
        assert cell.fill.fgColor.rgb.endswith("C6EFCE")

    def test_nearly_fit_row_is_yellow(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([_make_opp("nearly_fit")], path)
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Opportunities"]
        cell = ws.cell(row=2, column=1)
        assert cell.fill.fgColor.rgb.endswith("FFEB9C")

    def test_far_fetched_row_is_red(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([_make_opp("far_fetched")], path)
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Opportunities"]
        cell = ws.cell(row=2, column=1)
        assert cell.fill.fgColor.rgb.endswith("FFC7CE")

    def test_empty_list_only_has_header_row(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([], path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Opportunities"]
        # max_row may be 1 (header only) or None
        assert (ws.max_row or 1) == 1

    def test_summary_counts_are_correct(self, tmp_output):
        path = tmp_output / "report.xlsx"
        opps = [
            _make_opp("fit",          "Opp A"),
            _make_opp("fit",          "Opp B"),
            _make_opp("nearly_fit",   "Opp C"),
            _make_opp("far_fetched",  "Opp D"),
        ]
        m.build_excel(opps, path)
        wb  = openpyxl.load_workbook(path)
        ws2 = wb["Summary"]

        rows = {row[0].value: row[1].value for row in ws2.iter_rows() if row[0].value}
        assert rows["Total Opportunities"] == 4
        assert rows["Strong Fit"]          == 2
        assert rows["Nearly Fit"]          == 1
        assert rows["Far Fetched"]         == 1

    def test_sectors_joined_as_string(self, tmp_output):
        path = tmp_output / "report.xlsx"
        opp  = _make_opp("fit")
        opp["sectors"] = ["Finance", "Data", "AI"]
        m.build_excel([opp], path)
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Opportunities"]
        # Sectors is column 4
        assert ws.cell(row=2, column=4).value == "Finance, Data, AI"

    def test_eligibility_bulleted(self, tmp_output):
        path = tmp_output / "report.xlsx"
        m.build_excel([_make_opp("fit")], path)
        wb   = openpyxl.load_workbook(path)
        ws   = wb["Opportunities"]
        # Eligibility is column 10
        cell_val = ws.cell(row=2, column=10).value
        assert "• Registered firm" in cell_val
